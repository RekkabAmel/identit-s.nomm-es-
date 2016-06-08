# -*- coding: utf-8 -*-
"""
Created on Tue Jun 7 14:35:58 2016

@author: daphnehb
"""
# pour les vieilles versions (Excel avant 2010)
from xlwt import Workbook as Old_Workbook
# pour versions récentes (Excel apres 2010)
from openpyxl import Workbook, cell as Xls_cell
import exceptions as exc
import os, subprocess
import string


def csv_to_xls_vOLD(csv_filename, xls_filename=None, separateur=";") :
    """
    Transformation d'un fichier CSV, avec sep comme séparation, en un fichier xsl (Excel)
    Retourne le nom du fichier xls et son arborescence
    """
    # on vérifie que fichier d'entrée est bien un fichier CSV
    if not csv_filename.lower().endswith(".csv"):
        raise exc.NotCSVFileException(csv_filename)

    if xls_filename is None:
        # TODO demander à l'utilisateur ce qu'il veut donner comme nom
        # on créé un nouveau nom
        pre, ext = os.path.splitext(csv_filename)
        xls_filename = str(pre) + ".xls"

    with open(csv_filename,"r") as csv_file:
        # on crée l'objet qui fera le fichier Excel
        book = Old_Workbook()
        # TODO : une feuille par facture (=> une liste de CSV en parametre)?
        feuille = book.add_sheet('Feuille 1')
        # ajout des en-têtes
        """feuille.write(0,0,'id')
        feuille.write(0,1,'x')
        feuille.write(0,2,'y')
        feuille.write(0,3,'test')"""
        # pour chaque ligne du fichier d'entrée on met les valeurs dans le fichier excel
        for num_ligne,ligne in enumerate(csv_file.readlines()):
            # recuperation de la reference à la ligne du fichier Excel
            ref_ligne = feuille.row(num_ligne)
            # on recupere la ligne-liste sans \n final, et séparant les cases selon le séparateur
            ligne = ligne.rstrip("\n").split(separateur)
            # remplisage de la ligne cellule par cellule
            for num_cell,cell in enumerate(ligne):
                ref_ligne.write(num_cell,str(cell))

        #feuille.col(0).width = 10000
        book.save(xls_filename)
        return xls_filename

def csv_to_xlsx(csv_filename, xlsx_filename=None, separateur=";") :
    """
    Version pour
    Transformation d'un fichier CSV, avec separateur comme séparation, en un fichier xsl (Excel)
    Retourne le nom du fichier xls et son arborescence
    """
    # on vérifie que fichier d'entrée est bien un fichier CSV
    if not csv_filename.lower().endswith(".csv"):
        raise exc.NotCSVFileException(csv_filename)

    if xlsx_filename is None:
        # TODO demander à l'utilisateur ce qu'il veut donner comme nom
        # on créé un nouveau nom
        pre, ext = os.path.splitext(csv_filename)
        xlsx_filename = str(pre) + ".xlsx"

    #excel_columns = list(string.ascii_uppercase)
    with open(csv_filename,"r") as csv_file:
        # on crée l'objet qui fera le fichier Excel
        book = Workbook()
        # TODO : une feuille par facture (=> une liste de CSV en parametre)?
        # for feuille in book.get_sheet_names()
        feuille = book.active
        feuille.title = "Feuille 1"
        # ajout des en-têtes
        """feuille.write(0,0,'id')
        feuille.write(0,1,'x')
        feuille.write(0,2,'y')
        feuille.write(0,3,'test')"""

        # pour chaque ligne du fichier d'entrée on met les valeurs dans le fichier excel
        for num_ligne,ligne in enumerate(csv_file.readlines(), start=1):
            # recuperation de la reference à la ligne du fichier Excel
            #ref_ligne = feuille.row(num_ligne)
            # on recupere la ligne-liste sans \n final, et séparant les cases selon le séparateur
            ligne = ligne.rstrip("\n").split(separateur)
            # remplisage de la ligne cellule par cellule
            for num_cell,cell in enumerate(ligne, start=1):
                col_lettre = Xls_cell.get_column_letter(num_cell)
                num_case = str(col_lettre) + str(num_ligne)
                feuille[num_case] = str(cell)

        #feuille.col(0).width = 10000
        book.save(xlsx_filename)
        return xlsx_filename


############ TESTS
csv_to_xlsx("csv_sample.csv", "xls_sample.xlsx")
csv_to_xls_vOLD("csv_sample.csv", "xls_sample.xls")

excelProcess = subprocess.Popen(["libreoffice","xls_sample.xlsx"])
# on windows : excelProcess = popen2.Popen4("start excel %s" % (excelFile))
"""from win32com.client import Dispatch

xl = Dispatch("Excel.Application")
xl.Visible = True # otherwise excel is hidden

# newest excel does not accept forward slash in path
wb = xl.Workbooks.Open(r"xls_sample.xls")
wb.Close()
xl.Quit()
"""
