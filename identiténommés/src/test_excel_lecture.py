# -*- coding: utf-8 -*-
"""
Created on Tue Jun 7 14:35:58 2016

@author: daphnehb
"""

# pour les vieilles versions (Excel avant 2010)
from xlrd import open_workbook
# pour versions récentes (Excel apres 2010)
from openpyxl import Workbook, cell as Xls_cell
import exceptions as exc
import os, subprocess
import string


def xls_to_csv(xls_filename, csv_filename=None, separateur=";") :
    """
    Transformation d'un fichier Excel, avec separateur comme séparation, en un fichier CSV
    Retourne le nom du fichier CSV et son arborescence
    """
    csv_string = ""
    # dans le cas où le fichier d'entrée est un fichier XLS (Excel avant 2010)
    if xls_filename.lower().endswith(".xls"):
        # ouverture du fichier Excel
        book = open_workbook(xls_filename)

        # lecture des données dans la première feuille
        sh = book.sheet_by_name(u'Feuille 1')
        for rownum in range(sh.nrows):
            print sh.row_values(rownum)

        # lecture par colonne
        colonne1 = sh.col_values(0)
        print colonne1
        #[u'id', 1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0]

        colonne2=sh.col_values(1)
        print colonne2
        #[u'x', 235.0, 245.0, 255.0, 265.0, 275.0, 285.0, 295.0]

        # extraction d'un élément particulier
        print colonne1[1],colonne2[1]
    # dans le cas où le fichier d'entrée est un fichier XLSX (Excel après 2010)
    elif xls_filename.lower().endswith(".xlsx"):
        book = openpyxl.load_workbook(xls_filename)
        # TODO multiple sheets ?
        feuille = book.active
        for ligne in range(feuille.max_row):
            for cell in range(feuille.max_column):
                csv_string += str(separateur)+" "
            # on retire les deux derniers caractères (separateur et espace)
            csv_string = csv_string[0:-2]
            csv_string += "\n"
    else
        # TODO raise not an Excel file
        pass
        return None
    # TODO save csv file
    return csv_filename
